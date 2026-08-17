"""
============================================================
APLIKASI MONITORING TRANSMITTER
DENGAN DETEKSI ANOMALI ISOLATION FOREST
============================================================
"""

import streamlit as st
import pandas as pd
import numpy as np
import joblib
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, timedelta
import warnings
import os
warnings.filterwarnings('ignore')

# Import dari folder utils
from utils.normal_ranges import NORMAL_RANGES, PARAM_DISPLAY, UNITS
from utils.helpers import get_status, get_global_status
from utils.data_loader import load_clean_data, load_preprocessed_data, load_model, get_combined_data

# ============================================================
# KONFIGURASI HALAMAN
# ============================================================
st.set_page_config(
    page_title="Monitoring Transmitter",
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# SISTEM LOGIN
# ============================================================

# Inisialisasi session state untuk login
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "username" not in st.session_state:
    st.session_state.username = ""

# File untuk menyimpan data user
USER_FILE = 'data/users.csv'

def load_users():
    """Load user dari CSV"""
    if os.path.exists(USER_FILE):
        df = pd.read_csv(USER_FILE)
        return dict(zip(df['username'], df['password']))
    else:
        # Default jika file belum ada
        return {"operator": "transmitter123", "admin": "admin123"}

def save_users(users_dict):
    """Simpan user ke CSV"""
    df = pd.DataFrame(list(users_dict.items()), columns=['username', 'password'])
    df.to_csv(USER_FILE, index=False)

# Load user dari file
USER_CREDENTIALS = load_users()

def login_page():
    """Halaman login di tengah"""
    
    # Layout 3 kolom, konten di tengah
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        # Logo di tengah dengan ukuran yang pas
        col_logo1, col_logo2, col_logo3 = st.columns([1, 2, 1])
        with col_logo2:
            st.image("newest_transmonitor.png", width=300) 
        
        st.markdown("<br>", unsafe_allow_html=True)  # Spasi antara logo dan form login
        
        # Form Login
        st.subheader("🔐 Login")
        username = st.text_input("Username", placeholder="Masukkan username")
        password = st.text_input("Password", type="password", placeholder="Masukkan password")
        
        # Tombol Login saja (tanpa Reset)
        if st.button("Login", use_container_width=True, type="primary"):
            # Reload user dari file (update terbaru)
            users = load_users()
            if username in users and users[username] == password:
                st.session_state.logged_in = True
                st.session_state.username = username
                st.rerun()
            else:
                st.error("❌ Username atau password salah!")
        
        # --- FITUR RESET PASSWORD ---
        st.markdown("---")
        with st.expander("🔑 Lupa Password? Reset Sekarang!"):
            st.warning("⚠️ Reset password akan mengubah password Anda.")
            
            reset_username = st.text_input("Username Anda", placeholder="Masukkan username", key="reset_user")
            new_password = st.text_input("Password Baru (min 6 karakter)", type="password", placeholder="Masukkan password baru", key="new_pass")
            confirm_password = st.text_input("Konfirmasi Password Baru", type="password", placeholder="Ulangi password baru", key="confirm_pass")
            
            if st.button("Reset Password", use_container_width=True, key="reset_btn"):
                users = load_users()
                if reset_username in users:
                    if new_password == confirm_password and len(new_password) >= 6:
                        # Update password
                        users[reset_username] = new_password
                        save_users(users)
                        st.balloons()  # Animasi balon
                        st.success(f"✅ Password untuk user '{reset_username}' BERHASIL direset!")
                        st.info("🔐 Silakan login dengan password baru Anda.")
                        st.rerun()
                    elif len(new_password) < 6:
                        st.error("❌ Password minimal 6 karakter!")
                    else:
                        st.error("❌ Password dan konfirmasi tidak sama!")
                else:
                    st.error("❌ Username tidak ditemukan!")
        
        st.caption("© 2026 TransMonitor")

def logout():
    """Fungsi logout"""
    st.session_state.logged_in = False
    st.session_state.username = ""
    st.rerun()

# Jika belum login, tampilkan halaman login di tengah
if not st.session_state.logged_in:
    login_page()
    st.stop()

# ============================================================
# LOAD DATA DAN MODEL
# ============================================================

# Load data
df_clean = load_clean_data()
df_prep = load_preprocessed_data()
model, threshold = load_model()

# Buat data gabungan (utama + input) dengan prediksi
df_combined = get_combined_data(df_clean, model)

# Feature names
feature_names = ['Vision Output Power (KW)', 'Beam Voltage (KV)', 'Beam Current (A)', 
                 'Driver FWD Power (W)', 'Water Temp In (C)', 'Water Temp Out (C)']

# ============================================================
# UPDATE PREDIKSI jika perlu (dari data preprocessed)
# ============================================================
if 'prediction' in df_prep.columns and df_prep['prediction'].nunique() == 1:
    try:
        X = df_prep[feature_names].values
        df_prep['prediction'] = model.predict(X)
        df_prep['anomaly_score'] = model.decision_function(X)
        df_prep['anomaly_status'] = df_prep['prediction'].map({1: 'Normal', -1: 'Anomali'})
    except Exception as e:
        pass


# ============================================================
# SIDEBAR NAVIGASI
# ============================================================
st.sidebar.image("newest_transmonitor.png", width=380)
st.sidebar.markdown("---")

# Tampilkan user yang login
st.sidebar.success(f"👤 Login sebagai: {st.session_state.username}")
if st.sidebar.button("🚪 Logout", use_container_width=True):
    logout()

st.sidebar.markdown("---")

menu = st.sidebar.radio(
    "Pilih Menu",
    ["🏠 Dashboard", "📥 Input Data", "📋 Riwayat", "⬇️ Ekspor"],
    index=0
)

st.sidebar.markdown("---")

# ============================================================
# HALAMAN 1: DASHBOARD
# ============================================================
if menu == "🏠 Dashboard":
    st.title("🏠 Dashboard Monitoring Transmitter")
    
    # Ambil data gabungan untuk dashboard
    df_combined = get_combined_data(df_clean, model)
    
    # Ambil data terakhir dari data GABUNGAN
    latest_data = df_combined.iloc[-1]
    
    # --- KARTU INDIKATOR ---
    st.subheader("📊 Status Parameter Terkini")
    
    cols = st.columns(3)
    
    for idx, (col_name, ranges) in enumerate(NORMAL_RANGES.items()):
        with cols[idx % 3]:
            value = latest_data[col_name]
            status, icon = get_status(value, col_name)
            unit = ranges.get('unit', '')
            decimals = ranges.get('decimals', 2)
            
            if status == 'danger':
                bg_color = '#ffebee'
                border_color = '#c62828'
                status_text = '⚠️ DANGER'
            elif status == 'warning':
                bg_color = '#fff3e0'
                border_color = '#ef6c00'
                status_text = '⚠️ WARNING'
            else:
                bg_color = '#e8f5e9'
                border_color = '#2e7d32'
                status_text = '✅ NORMAL'
            
            st.markdown(f"""
            <div style="
                background-color: {bg_color};
                border-left: 5px solid {border_color};
                padding: 12px 16px;
                border-radius: 10px;
                margin-bottom: 12px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            ">
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <span style="font-size: 14px; font-weight: 600; color: #333;">{PARAM_DISPLAY[col_name]}</span>
                    <span style="font-size: 20px;">{icon}</span>
                </div>
                <div style="font-size: 24px; font-weight: bold; color: #222; margin: 4px 0;">
                    {value:.{decimals}f} {unit}
                </div>
                <div style="font-size: 12px; font-weight: 500; color: {border_color};">
                    {status_text}
                </div>
                <div style="font-size: 11px; color: #888; margin-top: 2px;">
                    Normal: {ranges['min']} - {ranges['max']} {unit}
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    st.markdown("---")
    
            # --- GRAFIK TREN ---
    st.subheader("📈 Grafik Tren Parameter")
    
    col_display1, col_display2 = st.columns([2, 1])
    
    with col_display1:
        selected_params = st.multiselect(
            "Pilih Parameter",
            list(NORMAL_RANGES.keys()),
            default=list(NORMAL_RANGES.keys())[:3]
        )
    
    with col_display2:
        n_options = [
            {"label": "6 jam terakhir", "value": 6},
            {"label": "12 jam terakhir", "value": 12},
            {"label": "1 hari terakhir", "value": 24},
            {"label": "2 hari terakhir", "value": 48},
            {"label": "3 hari terakhir", "value": 72},
        ]
        n_last = st.selectbox(
            "📊 Tampilkan data terakhir",
            options=n_options,
            format_func=lambda x: x["label"],
            index=2
        )
    
    if selected_params:
        n_data = n_last["value"] // 2
        if n_data > len(df_combined):
            n_data = len(df_combined)
        
        plot_df = df_combined.tail(int(n_data)).copy()
        
        plot_df['datetime'] = pd.to_datetime(
            plot_df['Tanggal'].astype(str) + ' ' + plot_df['Jam'].astype(str),
            errors='coerce'
        )
        plot_df = plot_df.dropna(subset=['datetime'])
        plot_df = plot_df.sort_values('datetime').reset_index(drop=True)
        
        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b']
        
        fig = go.Figure()
        
        for idx, param in enumerate(selected_params):
            color = colors[idx % len(colors)]
            
            # TAMBAHKAN SEMUA DATA (tanpa membedakan normal/anomali)
            fig.add_trace(go.Scatter(
                x=plot_df['datetime'],
                y=plot_df[param],
                mode='lines+markers',
                name=PARAM_DISPLAY[param],
                marker=dict(size=6, color=color, symbol='circle'),
                line=dict(width=2, color=color),
                hovertemplate=(
                    '<b>%{x|%d/%m/%Y %H:%M}</b><br>' +
                    f'{PARAM_DISPLAY[param]}: %{{y:.2f}}<br>' +
                    '<extra></extra>'
                )
            ))
        
        fig.update_layout(
            height=450,
            title="Perubahan Parameter dari Waktu ke Waktu",
            xaxis_title='Waktu',
            yaxis_title='Nilai Parameter',
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
            hovermode='x unified',
            xaxis=dict(
                title=dict(text='Waktu', font=dict(size=12, color='#666')),
                tickformat='%H:%M<br>%d/%m/%Y',  # 00:00 di atas, 25/07/2026 di bawah
                tickangle=0,
                tickmode='array',
                tickvals=plot_df['datetime'].tolist(),
                ticktext=[d.strftime('%H:%M<br>%d/%m/%Y') for d in plot_df['datetime']],
                nticks=min(len(plot_df), 15),
                showgrid=True,
                gridcolor='#f0f0f0'
            ),
            yaxis=dict(
                showgrid=True,
                gridcolor='#f0f0f0'
            ),
            plot_bgcolor='white'
        )
        
        st.plotly_chart(fig, use_container_width=True)
        
        col_info1, col_info2, col_info3 = st.columns(3)
        with col_info1:
            st.caption(f"📊 Menampilkan {len(plot_df)} data terakhir ({n_last['label']})")
        with col_info2:
            st.caption(f"📅 Total data: {len(df_combined)} record")
        with col_info3:
            st.caption(f"🕐 Frekuensi: setiap 2 jam (12x/hari)")
        
    else:
        st.info("👆 Pilih minimal satu parameter untuk ditampilkan")
    
    st.markdown("---")
    
    # --- ALERT ---
    st.subheader("⚠️ Catatan")
    
    alerts = []
    for col in NORMAL_RANGES.keys():
        if col in latest_data:
            value = latest_data[col]
            status, _ = get_status(value, col)
            if status == 'danger':
                alerts.append(f"🚨 {PARAM_DISPLAY[col]}: {value:.2f} (melewati batas normal!)")
            elif status == 'warning':
                alerts.append(f"🟡 {PARAM_DISPLAY[col]}: {value:.2f} (mendekati batas)")
    
    if alerts:
        for alert in alerts:
            if '🚨' in alert:
                st.error(alert)
            else:
                st.warning(alert)
    else:
        st.success("✅ Tidak ada anomali terdeteksi. Semua parameter normal.")

# ============================================================
# HALAMAN 2: INPUT DATA
# ============================================================
elif menu == "📥 Input Data":
    st.title("📥 Input Data Parameter Transmitter")
    st.markdown("---")
    
    st.info("📝 Masukkan data parameter transmitter untuk dilakukan deteksi anomali")
    
    # File untuk menyimpan data input
    DATA_FILE = 'data/transmitter_data_input.csv'
    
    # Fungsi untuk menyimpan data ke CSV
    def save_input_data(tanggal, jam, values):
        """Simpan data input ke file CSV"""
        
        # Buat dictionary data
        new_data = {
            'Tanggal': [tanggal],
            'Jam': [jam],
            'Vision Output Power (KW)': [values[0]],
            'Beam Voltage (KV)': [values[1]],
            'Beam Current (A)': [values[2]],
            'Driver FWD Power (W)': [values[3]],
            'Water Temp In (C)': [values[4]],
            'Water Temp Out (C)': [values[5]]
        }
        
        df_new = pd.DataFrame(new_data)
        
        # Cek apakah file sudah ada
        if os.path.exists(DATA_FILE):
            # Baca data lama
            df_existing = pd.read_csv(DATA_FILE)
            # Gabungkan dengan data baru
            df_combined = pd.concat([df_existing, df_new], ignore_index=True)
        else:
            df_combined = df_new
        
        # Simpan ke CSV
        df_combined.to_csv(DATA_FILE, index=False)
        return True
    
    with st.form("input_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        
        with col1:
            tanggal = st.date_input("📅 Tanggal", datetime.now())
            jam_options = [f"{h:02d}:00" for h in range(0, 24, 2)]
            jam = st.selectbox("⏰ Jam", jam_options, index=datetime.now().hour // 2)
        
        with col2:
            st.write("**📡 Parameter Transmitter**")
            vision_output = st.number_input("Vision Output Power (KW)", min_value=0.0, value=14.5, step=0.1, format="%.2f")
            beam_voltage = st.number_input("Beam Voltage (KV)", min_value=0.0, value=31.0, step=0.1, format="%.2f")
            beam_current = st.number_input("Beam Current (A)", min_value=0.0, value=1.1, step=0.01, format="%.2f")
            driver_forward = st.number_input("Driver FWD Power (W)", min_value=0.0, value=188.0, step=1.0, format="%.0f")
            temp_in = st.number_input("Water Temp In (°C)", min_value=0.0, value=40.0, step=1.0, format="%.0f")
            temp_out = st.number_input("Water Temp Out (°C)", min_value=0.0, value=48.0, step=1.0, format="%.0f")
        
        st.markdown("---")
        
        # Tombol Simpan
        submitted = st.form_submit_button("💾 Simpan", use_container_width=True, type="primary")
    
    if submitted:
        values = [vision_output, beam_voltage, beam_current, driver_forward, temp_in, temp_out]
        
        # Validasi: cek apakah ada nilai negatif
        if any(v < 0 for v in values):
            st.error("❌ Nilai tidak boleh negatif!")
        else:
            # Simpan data
            try:
                tanggal_str = tanggal.strftime('%Y-%m-%d')
                save_input_data(tanggal_str, jam, values)
                st.success("✅ Data berhasil disimpan ke database!")
                
                # Tampilkan hasil deteksi status
                st.subheader("📊 Hasil Deteksi Status")
                
                cols = st.columns(3)
                param_names = list(NORMAL_RANGES.keys())
                
                for idx, col_name in enumerate(param_names):
                    with cols[idx % 3]:
                        value = values[idx]
                        status, icon_status = get_status(value, col_name)
                        
                        if status == 'danger':
                            bg_color = '#ffebee'
                            status_text = '⚠️ DANGER'
                        elif status == 'warning':
                            bg_color = '#fff3e0'
                            status_text = '⚠️ WARNING'
                        else:
                            bg_color = '#e8f5e9'
                            status_text = '✅ NORMAL'
                        
                        st.markdown(f"""
                        <div style="
                            background-color: {bg_color};
                            padding: 12px;
                            border-radius: 8px;
                            margin-bottom: 10px;
                            border-left: 5px solid {'#c62828' if status == 'danger' else '#ef6c00' if status == 'warning' else '#2e7d32'};
                        ">
                            <div style="display: flex; justify-content: space-between;">
                                <span style="font-size: 13px; color: #666;">{PARAM_DISPLAY[col_name]}</span>
                                <span style="font-size: 16px;">{icon_status}</span>
                            </div>
                            <div style="font-size: 20px; font-weight: bold;">
                                {value:.2f} {UNITS[col_name]}
                            </div>
                            <div style="font-size: 11px; color: #666;">
                                {status_text}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                
                # Tampilkan status global
                if any(get_status(v, param_names[i])[0] == 'danger' for i, v in enumerate(values)):
                    st.error("🚨 Terdapat parameter dalam kondisi DANGER! Segera periksa perangkat!")
                elif any(get_status(v, param_names[i])[0] == 'warning' for i, v in enumerate(values)):
                    st.warning("⚠️ Terdapat parameter dalam kondisi WARNING. Perlu perhatian!")
                else:
                    st.success("✅ Semua parameter dalam kondisi NORMAL.")
                
            except Exception as e:
                st.error(f"❌ Gagal menyimpan data: {e}")

# ============================================================
# HALAMAN 3: RIWAYAT
# ============================================================
elif menu == "📋 Riwayat":
    st.title("📋 Riwayat Data Parameter")
    st.markdown("---")
    
    # Ambil data gabungan untuk riwayat
    df_combined = get_combined_data(df_clean, model)
    
    col_filter1 = st.columns(1)[0]
    
    with col_filter1:
        min_date = pd.to_datetime(df_combined['Tanggal'].min()).date()
        max_date = pd.to_datetime(df_combined['Tanggal'].max()).date()
        date_range = st.date_input(
            "📅 Pilih Rentang Tanggal",
            [min_date, max_date],
            min_value=min_date,
            max_value=max_date
        )
    
    # Gunakan data gabungan
    filtered_df = df_combined.copy()
    
    if len(date_range) == 2:
        start_date, end_date = date_range
        filtered_df = filtered_df[
            (pd.to_datetime(filtered_df['Tanggal']).dt.date >= start_date) &
            (pd.to_datetime(filtered_df['Tanggal']).dt.date <= end_date)
        ]
    
    st.subheader(f"📊 Data Riwayat ({len(filtered_df)} record)")
    
    if len(filtered_df) > 0:
        base_cols = ['Tanggal', 'Jam']
        param_cols = list(NORMAL_RANGES.keys())
        
        display_cols = base_cols + param_cols
        display_df = filtered_df[display_cols].copy()
        
        # Tampilkan data dengan checkbox untuk hapus per baris
        st.dataframe(
            display_df,
            use_container_width=True,
            height=400
        )
        
                # ============================================================
        # FITUR HAPUS PER BARIS (DENGAN FILTER TANGGAL)
        # ============================================================
        st.markdown("---")
        st.subheader("🗑️ Hapus Data Input (Per Baris)")
        
        # Tampilkan data input yang bisa dihapus
        input_file = 'data/transmitter_data_input.csv'
        if os.path.exists(input_file):
            df_input = pd.read_csv(input_file)
            if len(df_input) > 0:
                st.info(f"📥 Total data input: {len(df_input)} data")
                
                # --- FILTER TANGGAL UNTUK HAPUS ---
                st.write("**🔍 Filter data yang ingin dihapus:**")
                
                col_filter_tanggal1, col_filter_tanggal2 = st.columns(2)
                
                with col_filter_tanggal1:
                    # Ambil tanggal min dan max dari data input
                    min_date_input = pd.to_datetime(df_input['Tanggal'].min()).date()
                    max_date_input = pd.to_datetime(df_input['Tanggal'].max()).date()
                    
                    start_date_filter = st.date_input(
                        "Dari tanggal",
                        min_date_input,
                        min_value=min_date_input,
                        max_value=max_date_input,
                        key="start_date_delete"
                    )
                
                with col_filter_tanggal2:
                    end_date_filter = st.date_input(
                        "Sampai tanggal",
                        max_date_input,
                        min_value=min_date_input,
                        max_value=max_date_input,
                        key="end_date_delete"
                    )
                
                # Filter data input berdasarkan tanggal
                if start_date_filter and end_date_filter:
                    df_input_filtered = df_input[
                        (pd.to_datetime(df_input['Tanggal']).dt.date >= start_date_filter) &
                        (pd.to_datetime(df_input['Tanggal']).dt.date <= end_date_filter)
                    ]
                else:
                    df_input_filtered = df_input
                
                if len(df_input_filtered) == 0:
                    st.info("Tidak ada data input pada rentang tanggal yang dipilih.")
                else:
                    st.write(f"**📋 Menampilkan {len(df_input_filtered)} data dari {len(df_input)} total data input**")
                    
                    # Tampilkan data input dengan checkbox
                    st.write("**Pilih data yang ingin dihapus:**")
                    
                    # Buat checkbox untuk setiap baris data input
                    selected_indices = []
                    
                    # Tampilkan dalam format yang lebih ringkas (2 kolom)
                    for i, (idx, row) in enumerate(df_input_filtered.iterrows()):
                        # Gunakan 2 kolom: checkbox + info singkat
                        col_check, col_info = st.columns([1, 6])
                        
                        with col_check:
                            checked = st.checkbox("", key=f"del_{idx}")
                        
                        with col_info:
                            st.write(f"📅 {row['Tanggal']} ⏰ {row['Jam']} | "
                                    f"V.Output: {row['Vision Output Power (KW)']:.2f} KW | "
                                    f"Beam V: {row['Beam Voltage (KV)']:.2f} KV")
                        
                        if checked:
                            selected_indices.append(idx)
                    
                    # Tombol hapus yang dipilih
                    if selected_indices:
                        st.markdown("---")
                        col_btn1, col_btn2 = st.columns(2)
                        with col_btn1:
                            if st.button(f"🗑️ Hapus {len(selected_indices)} Data Terpilih", 
                                        use_container_width=True, type="primary"):
                                try:
                                    # Hapus data yang dipilih
                                    df_input_updated = df_input.drop(selected_indices).reset_index(drop=True)
                                    df_input_updated.to_csv(input_file, index=False)
                                    st.success(f"✅ {len(selected_indices)} data berhasil dihapus!")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ Gagal menghapus data: {e}")
                        with col_btn2:
                            if st.button("❌ Batal Pilih", use_container_width=True):
                                st.rerun()
                    else:
                        st.caption("💡 Centang data yang ingin dihapus, lalu klik tombol hapus.")
                    
                    # Tombol pilih semua dalam rentang tanggal
                    col_select_all1, col_select_all2 = st.columns(2)
                    with col_select_all1:
                        if st.button("✅ Pilih Semua (rentang tanggal ini)", use_container_width=True):
                            # Pilih semua data dalam rentang tanggal
                            for idx in df_input_filtered.index:
                                # Set session state untuk checkbox
                                st.session_state[f"del_{idx}"] = True
                            st.rerun()
                    with col_select_all2:
                        if st.button("⬜ Hapus Pilihan", use_container_width=True):
                            for idx in df_input_filtered.index:
                                st.session_state[f"del_{idx}"] = False
                            st.rerun()
            else:
                st.info("Tidak ada data input yang tersimpan.")
        else:
            st.info("Tidak ada data input yang tersimpan.")
        
        # ============================================================
        # TOMBOL HAPUS SEMUA (Opsional)
        # ============================================================
        st.markdown("---")
        st.subheader("⚠️ Hapus Semua Data Input")
        
        col_del1, col_del2 = st.columns([3, 1])
        with col_del1:
            st.caption("⚠️ Menghapus **semua** data input. Data utama (Maret-April 2025) tetap aman.")
        with col_del2:
            if st.button("🗑️ Hapus Semua", use_container_width=True, type="secondary"):
                st.session_state.confirm_delete_all = True
        
        if st.session_state.get("confirm_delete_all", False):
            st.error("⚠️ **Konfirmasi Hapus Semua**")
            st.warning("Anda yakin ingin menghapus **semua** data input? Tindakan ini tidak dapat dibatalkan!")
            
            col_confirm1, col_confirm2 = st.columns(2)
            with col_confirm1:
                if st.button("✅ Ya, Hapus Semua", use_container_width=True, type="primary"):
                    try:
                        if os.path.exists(input_file):
                            os.remove(input_file)
                            st.success("✅ Semua data input berhasil dihapus!")
                            st.session_state.confirm_delete_all = False
                            st.rerun()
                        else:
                            st.info("Tidak ada data input yang ditemukan.")
                    except Exception as e:
                        st.error(f"❌ Gagal menghapus data: {e}")
            
            with col_confirm2:
                if st.button("❌ Batal", use_container_width=True):
                    st.session_state.confirm_delete_all = False
                    st.rerun()
        
    else:
        st.info("Tidak ada data yang sesuai dengan filter")

# ============================================================
# HALAMAN 4: EKSPOR
# ============================================================
elif menu == "⬇️ Ekspor":
    st.title("⬇️ Ekspor Data Laporan")
    st.markdown("---")
    
    # Ambil data gabungan untuk ekspor
    df_combined = get_combined_data(df_clean, model)
    
    st.subheader("📅 Pilih Periode Laporan")
    
    col_period1, col_period2 = st.columns([2, 3])
    
    with col_period1:
        period = st.radio(
            "Periode",
            ["Hari Ini", "Minggu Ini", "Bulan Ini", "Kustom"],
            index=0,
            horizontal=True
        )
    
    with col_period2:
        if period == "Hari Ini":
            start_date = datetime.now().date()
            end_date = datetime.now().date()
            st.write(f"📅 {start_date.strftime('%d/%m/%Y')}")
        elif period == "Minggu Ini":
            today = datetime.now().date()
            start_date = today - timedelta(days=today.weekday())
            end_date = today
            st.write(f"📅 {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}")
        elif period == "Bulan Ini":
            today = datetime.now().date()
            start_date = today.replace(day=1)
            end_date = today
            st.write(f"📅 {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}")
        else:
            col_custom1, col_custom2 = st.columns(2)
            with col_custom1:
                start_date = st.date_input("Dari", datetime.now().date() - timedelta(days=30))
            with col_custom2:
                end_date = st.date_input("Sampai", datetime.now().date())
    
    st.subheader("📄 Pilih Format File")
    
    col_format1, col_format2, col_format3 = st.columns(3)
    
    with col_format1:
        format_csv = st.checkbox("📄 CSV (Data Mentah)", value=True)
    with col_format2:
        format_excel = st.checkbox("📊 Excel (Data Terstruktur)", value=False)
    with col_format3:
        format_pdf = st.checkbox("📑 PDF (Laporan Resmi)", value=False)
    
    # --- PILIH PARAMETER YANG AKAN DIEKSPOR ---
    st.subheader("🔍 Pilih Parameter yang Diekspor")
    param_options = ["Semua Parameter"] + list(PARAM_DISPLAY.values())
    selected_param_export = st.radio(
        "Parameter",
        options=param_options,
        index=0,
        horizontal=True,
        format_func=lambda x: x
    )
    
    st.markdown("---")
    
    # Siapkan data untuk diekspor (gunakan data gabungan)
    export_df = df_combined.copy()
    
    export_df = export_df[
        (pd.to_datetime(export_df['Tanggal']).dt.date >= start_date) &
        (pd.to_datetime(export_df['Tanggal']).dt.date <= end_date)
    ]
    
    # --- TAMBAHKAN KOLOM STATUS UNTUK SETIAP PARAMETER ---
    for col in NORMAL_RANGES.keys():
        if col in export_df.columns:
            export_df[f'Status_{PARAM_DISPLAY[col]}'] = export_df[col].apply(lambda x: get_status(x, col)[0].upper())
    
    # Hapus kolom teknis (prediction, anomaly_score, anomaly_status) sebelum ekspor
    export_df = export_df.drop(columns=['prediction', 'anomaly_score', 'anomaly_status'], errors='ignore')
    
    # --- FILTER BERDASARKAN PARAMETER YANG DIPILIH ---
    export_columns = ['Tanggal', 'Jam']
    
    if selected_param_export == "Semua Parameter":
        # Tampilkan semua parameter
        for col in NORMAL_RANGES.keys():
            export_columns.append(col)
            export_columns.append(f'Status_{PARAM_DISPLAY[col]}')
    else:
        # Cari nama parameter asli dari display name
        for key, display in PARAM_DISPLAY.items():
            if display == selected_param_export:
                export_columns.append(key)
                export_columns.append(f'Status_{display}')
                break
    
    # Filter kolom yang akan diekspor
    export_df = export_df[export_columns]
    
    # Jika tidak ada data
    if len(export_df) == 0:
        st.warning("⚠️ Tidak ada data untuk periode yang dipilih!")
    
    # Tombol download utama
    if st.button("📥 DOWNLOAD LAPORAN", use_container_width=True, type="primary"):
        if len(export_df) == 0:
            st.warning("⚠️ Tidak ada data untuk periode yang dipilih!")
        else:
            # --- FORMAT CSV ---
            if format_csv:
                csv = export_df.to_csv(index=False)
                st.download_button(
                    label="📥 Download CSV (Data Mentah)",
                    data=csv,
                    file_name=f"laporan_transmitter_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            # --- FORMAT EXCEL (XLSX) ---
            if format_excel:
                try:
                    import io
                    
                    # Buat Excel dengan pandas ExcelWriter
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # Tulis data ke sheet
                        export_df.to_excel(writer, sheet_name='Data Transmitter', index=False)
                        
                        # Ambil workbook untuk styling
                        workbook = writer.book
                        worksheet = writer.sheets['Data Transmitter']
                        
                        # Style header
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                        
                        header_font = Font(size=11, bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                        header_alignment = Alignment(horizontal="center", vertical="center")
                        border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        # Apply style ke header (baris 1)
                        for col in range(1, len(export_df.columns) + 1):
                            cell = worksheet.cell(row=1, column=col)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = border
                        
                        # Auto-fit lebar kolom
                        for col in worksheet.columns:
                            max_length = 0
                            column_letter = col[0].column_letter
                            for cell in col:
                                try:
                                    if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 25)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    output.seek(0)
                    
                    st.download_button(
                        label="📥 Download Excel (Data Terstruktur)",
                        data=output,
                        file_name=f"laporan_transmitter_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                except ImportError:
                    st.error("❌ Library openpyxl tidak terinstall. Silakan install: pip install openpyxl")
                except Exception as e:
                    st.error(f"❌ Terjadi kesalahan: {e}")
            
            # --- FORMAT PDF ---
            if format_pdf:
                try:
                    from reportlab.lib.pagesizes import landscape, A3
                    from reportlab.platypus import (
                        SimpleDocTemplate,
                        Table,
                        TableStyle,
                        Paragraph,
                        Spacer,
                    )
                    from reportlab.lib import colors
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib.enums import TA_CENTER
                    from reportlab.lib.units import inch
                    import io

                    pdf_buffer = io.BytesIO()

                    doc = SimpleDocTemplate(
                        pdf_buffer,
                        pagesize=landscape(A3),
                        rightMargin=20,
                        leftMargin=20,
                        topMargin=30,
                        bottomMargin=30
                    )

                    styles = getSampleStyleSheet()

                    title_style = ParagraphStyle(
                        'TitleStyle',
                        parent=styles['Heading1'],
                        fontSize=16,
                        alignment=TA_CENTER,
                        textColor=colors.HexColor("#1F4E79"),
                        spaceAfter=10
                    )

                    subtitle_style = ParagraphStyle(
                        'SubtitleStyle',
                        parent=styles['Normal'],
                        fontSize=11,
                        alignment=TA_CENTER,
                        spaceAfter=5
                    )

                    header_style = ParagraphStyle(
                        'HeaderStyle',
                        parent=styles['BodyText'],
                        alignment=TA_CENTER,
                        fontName="Helvetica-Bold",
                        fontSize=7,
                        leading=8,
                        textColor=colors.white
                    )

                    body_style = ParagraphStyle(
                        'BodyStyle',
                        parent=styles['BodyText'],
                        alignment=TA_CENTER,
                        fontSize=7,
                        leading=8
                    )

                    elements = []

                    # ==========================
                    # JUDUL
                    # ==========================
                    elements.append(Paragraph("LAPORAN DATA TRANSMITTER", title_style))
                    elements.append(Paragraph(
                        f"Periode: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
                        subtitle_style
                    ))
                    elements.append(Paragraph(
                        f"Tanggal Cetak: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                        subtitle_style
                    ))
                    elements.append(Spacer(1, 12))

                    # ==========================
                    # DATA
                    # ==========================

                    pdf_data = export_df.head(100)

                    PDF_HEADERS = {
                        "V_Output_Power": "V.Output",
                        "Beam_Voltage": "Beam Volt",
                        "Beam_Current": "Beam Curr",
                        "Driver_FW": "Driver FW",
                        "Temp_In": "Temp In",
                        "Temp_Out": "Temp Out",
                    }

                    headers = []

                    for col in pdf_data.columns:

                        if col == "Tanggal":
                            headers.append(Paragraph("Tanggal", header_style))

                        elif col == "Jam":
                            headers.append(Paragraph("Jam", header_style))

                        elif col in PDF_HEADERS:
                            headers.append(Paragraph(PDF_HEADERS[col], header_style))

                        elif col.startswith("Status_"):
                            headers.append(Paragraph("Status", header_style))

                        else:
                            headers.append(Paragraph(col, header_style))

                    table_data = [headers]

                    for row in pdf_data.itertuples(index=False):

                        row_data = []

                        for i, col in enumerate(pdf_data.columns):

                            val = row[i]

                            if isinstance(val, float):
                                if col in NORMAL_RANGES:
                                    decimals = NORMAL_RANGES[col].get("decimals", 2)
                                    text = f"{val:.{decimals}f}"
                                else:
                                    text = str(val)
                            else:
                                text = str(val)

                            row_data.append(Paragraph(text, body_style))

                        table_data.append(row_data)

                    # ==========================
                    # LEBAR KOLOM
                    # ==========================

                    col_widths = []

                    for col in pdf_data.columns:

                        if col == "Tanggal":
                            col_widths.append(0.90 * inch)

                        elif col == "Jam":
                            col_widths.append(0.60 * inch)

                        elif col.startswith("Status_"):
                            col_widths.append(0.75 * inch)

                        else:
                            col_widths.append(0.70 * inch)

                    table = Table(
                        table_data,
                        colWidths=col_widths,
                        repeatRows=1
                    )

                    table.setStyle(TableStyle([

                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F4E79")),
                        ('TEXTCOLOR', (0,0), (-1,0), colors.white),

                        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),

                        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),

                        ('FONTSIZE', (0,0), (-1,0), 8),
                        ('FONTSIZE', (0,1), (-1,-1), 7),

                        ('TOPPADDING', (0,0), (-1,0), 10),
                        ('BOTTOMPADDING', (0,0), (-1,0), 10),

                        ('TOPPADDING', (0,1), (-1,-1), 4),
                        ('BOTTOMPADDING', (0,1), (-1,-1), 4),

                        ('GRID', (0,0), (-1,-1), 0.4, colors.grey),

                        ('ROWBACKGROUNDS', (0,1), (-1,-1),
                            [colors.white, colors.HexColor("#F5F8FA")])

                    ]))

                    elements.append(table)
                    elements.append(Spacer(1, 12))

                    footer_style = ParagraphStyle(
                        'Footer',
                        parent=styles['Normal'],
                        fontSize=9,
                        textColor=colors.grey
                    )

                    footer = f"Total Data : {len(export_df)} Record"

                    if len(export_df) > 100:
                        footer += " (Menampilkan 100 data pertama)"

                    elements.append(Paragraph(footer, footer_style))

                    doc.build(elements)

                    pdf_buffer.seek(0)

                    st.download_button(
                        label="📥 Download PDF (Laporan Resmi)",
                        data=pdf_buffer,
                        file_name=f"laporan_transmitter_{datetime.now().strftime('%Y%m%d')}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )

                except ImportError:
                    st.error("Install reportlab terlebih dahulu : pip install reportlab")

                except Exception as e:
                    st.error(f"Gagal membuat PDF : {e}")
    
    # Info tambahan
    st.markdown("---")
    st.caption(f"📊 Total data yang akan diekspor: {len(export_df)} record")
    st.caption(f"📅 Periode: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}")
    
    if len(export_df) > 0:
        with st.expander("📋 Preview Data yang Akan Diekspor"):
            st.dataframe(export_df.head(10))